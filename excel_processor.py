import sys
from io import BytesIO
from pathlib import Path
from typing import List

import pandas as pd
import xlrd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def is_empty_sheet_openpyxl(ws):
    """
    ws: openpyxl sheet
    """
    for row in ws.iter_rows(min_row=1, max_row=100):
        for cell in row:
            if cell.value not in [None, "", " "]:
                return False
    # print(ws.merged_cells.ranges, ws.page_setup.orientation, ws.print_area)
    if any([
        len(ws.merged_cells.ranges) > 0,
        ws.page_setup.orientation != "portrait",
        ws.print_area not in [None, '', ' '],
    ]):
        return False

    return True


def is_empty_sheet_xlrd(ws):
    """ws: xlrd sheet"""
    nrows = ws.nrows
    ncols = ws.ncols
    if nrows == 0 or ncols == 0:
        return True
    return False


class ExcelParser:
    """
    解析excel文件，支持.xls、.xlsx格式，填充合并单元格，将每个sheet转换成markdown格式的table 便于llm使用
    支持.csv
    """

    def parse(self, input_data: str | Path | bytes | BytesIO, verbose=True, output_format='markdown') -> List[str]:
        """
        支持文件路径,Path,文件的bytes以及BytesIO
        """
        if isinstance(input_data, (str, Path)):
            path = Path(input_data)
            if not path.exists():
                raise FileNotFoundError(f'{path} is not exists.')
            if not path.is_file():
                raise ValueError(f'{path} is not a file.')
            file_bytes = path.read_bytes()
        elif isinstance(input_data, bytes):
            file_bytes = input_data
        elif isinstance(input_data, BytesIO):
            input_data.seek(0)
            file_bytes = input_data.read()
        else:
            raise TypeError(f'{input_data} is not a valid type. 输入类型必须是路径、bytes 或 BytesIO')

        file_type = xlrd.inspect_format(content=file_bytes)

        parsers = {
            'xls': (self.parse_xlrd, xlrd.biffh.XLRDError, "无法解析 .xls 文件，可能已损坏"),
            'xlsx': (self.parse_openpyxl, KeyError, "无法解析 .xlsx 文件，可能已损坏"),
            'csv': (self.parse_csv, Exception, "无法解析 .csv 文件，可能已损坏")
        }
        parser, error_type, error_message = parsers.get(file_type, (
            None, Exception, "无法解析文件，未知格式，当前只支持.xls/.xlsx/.csv格式"))

        if parser is None:
            raise ValueError(error_message)
        try:
            parse_res = parser(file_bytes, verbose=verbose, output_format=output_format)
        except error_type as e:
            raise ValueError(error_message) from e
        return parse_res

    @staticmethod
    def parse_openpyxl(file_data: bytes, verbose=True, output_format='markdown') -> List[str]:
        """
        将表格解析转换成markdown格式,只支持xlsx格式
        """
        if not isinstance(file_data, bytes):
            raise ValueError("file_data must be bytes")
        file_bio = BytesIO(file_data)
        wb = load_workbook(file_bio, data_only=True)
        result = []
        with pd.ExcelFile(file_bio) as xls:
            for sheet_name in xls.sheet_names:
                # 获取当前sheet的合并单元格信息
                ws = wb[sheet_name]

                if is_empty_sheet_openpyxl(ws):
                    if verbose:
                        print(f"跳过空工作表: {sheet_name}", file=sys.stderr)
                    continue

                merge_map = {}  # 记录合并单元格value
                merge_cell = {}  # 记录被合并单元格坐标
                merge_info = {}  # 记录合并单元格起始位置以及跨度
                # 构建合并单元格值映射表
                for merge_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(merge_range.coord)
                    master_value = ws.cell(min_row, min_col).value
                    merge_info[(min_row, min_col)] = (
                        max_row - min_row + 1,
                        max_col - min_col + 1
                    )
                    # 为合并区域所有单元格记录主值
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            merge_map[(row, col)] = master_value
                            if (row, col) != (min_row, min_col):
                                merge_cell[(row, col)] = True  # 记录被合并的单元格

                try:
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=None).astype(str)
                except Exception as e:
                    print(f"读取工作表 {sheet_name} 失败: {str(e)}", file=sys.stderr)
                    continue

                # 有效性验证（防止全空数据）
                if df.map(lambda x: x.strip() if isinstance(x, str) else '').eq('').all().all():
                    if verbose:
                        print(f"跳过无效数据工作表: {sheet_name}", file=sys.stderr)
                    continue

                # 应用合并单元格值覆盖
                for (row, col), value in merge_map.items():
                    # pandas的行号从0开始，openpyxl从1开始
                    df_row = row - 1
                    df_col = col - 1
                    # 确保索引不越界
                    if df_row < df.shape[0] and df_col < df.shape[1]:
                        df.iat[df_row, df_col] = str(value)
                # ==================== 转markdown ====================
                if output_format == 'markdown':
                    md_table = df.to_markdown()
                    # 添加表格元数据
                    md_table = f"# Sheet: {sheet_name}\n{md_table}"
                    result.append(md_table)
                else:
                    soup = BeautifulSoup(features='html.parser')
                    table = soup.new_tag('table', border="1")

                    thead = soup.new_tag('thead')  # 表头有多少行不好去确定
                    # tbody = soup.new_tag('tbody')  # TODO:本来想表头和数据部分分开 但是发现不好确定表头行数

                    # max_header_rows = 3
                    max_header_rows = df.shape[0]
                    for row_idx in range(max_header_rows):
                        tr = soup.new_tag('tr')
                        col_idx = 0  # 列索引 (0-based)
                        while col_idx < df.shape[1]:
                            # 转换为 openpyxl 的 1-based 索引
                            cell_pos = (row_idx + 1, col_idx + 1)

                            if cell_pos in merge_info:
                                rowspan, colspan = merge_info[cell_pos]
                                cell_value = ws.cell(*cell_pos).value
                                th = soup.new_tag('th', rowspan=str(rowspan), colspan=str(colspan))
                                th.string = str(cell_value)
                                tr.append(th)
                                col_idx += colspan
                            else:
                                if cell_pos not in merge_cell:
                                    th = soup.new_tag('th')
                                    th.string = str(df.iat[row_idx, col_idx])
                                    tr.append(th)
                                col_idx += 1

                        thead.append(tr)
                    table.append(thead)
                    result.append(str(table.prettify()))
        return result

    @staticmethod
    def parse_xlrd(file_data: bytes, verbose=True, output_format='markdown') -> List[str]:  # .xls
        """
        将表格解析转换成markdown格式,只支持xls格式
        支持转换为html的表格格式输出 更好适应多级表头
        """
        if not isinstance(file_data, bytes):
            raise ValueError("file_data must be bytes")
        file_bio = BytesIO(file_data)

        wb = xlrd.open_workbook(file_contents=file_data)
        sheet_names = wb.sheet_names()

        result = []

        for sheet_name in sheet_names[:1]:
            ws = wb.sheet_by_name(sheet_name)

            if is_empty_sheet_xlrd(ws):
                # print(f"{file_path} {sheet_name} is empty")
                print(f"跳过空工作表: {sheet_name}", file=sys.stderr)
                continue

            merge_map = {}  # 记录合并单元格value
            merge_cell = {}  # 记录被合并单元格坐标
            merge_info = {}  # 记录合并单元格起始位置以及跨度
            for merged_cell in ws.merged_cells:
                # cell (rlo, clo) (the top left one) will carry the data
                # rlo, rhi, clo, chi = merged_cell
                min_row, max_row, min_col, max_col = merged_cell
                master_value = ws.cell_value(min_row, min_col)
                merge_info[(min_row, min_col)] = (
                    max_row - min_row + 1,
                    max_col - min_col + 1
                )
                for row in range(min_row, max_row):
                    for col in range(min_col, max_col):
                        merge_map[(row, col)] = master_value
                        if (row, col) != (min_row, min_col):
                            merge_cell[(row, col)] = True  # 记录被合并的单元格

            try:
                df = pd.read_excel(file_bio, sheet_name=sheet_name, header=None).astype(str)
            except Exception as e:
                print(f"读取工作表 {sheet_name} 失败: {str(e)}")
                continue

            # 有效性验证（防止全空数据）
            if df.map(lambda x: x.strip() if isinstance(x, str) else '').eq('').all().all():
                if verbose:
                    print(f"跳过无效数据工作表: {sheet_name}")
                continue

            # 应用合并单元格值覆盖
            for (row, col), value in merge_map.items():
                # pandas的行号从0开始，xlrd从0开始
                df_row = row
                df_col = col
                # 确保索引不越界
                if df_row < df.shape[0] and df_col < df.shape[1]:
                    df.iat[df_row, df_col] = str(value)

            # ==================== 转markdown ====================
            if output_format == 'markdown':
                md_table = df.to_markdown()
                # 添加表格元数据
                md_table = f"# Sheet: {sheet_name}\n{md_table}"
                result.append(md_table)
            else:
                soup = BeautifulSoup(features='html.parser')
                table = soup.new_tag('table', border="1")

                thead = soup.new_tag('thead')  # 表头有多少行不好去确定
                # tbody = soup.new_tag('tbody')  # TODO:本来想表头和数据部分分开 但是发现不好确定表头行数

                # max_header_rows = 3
                max_header_rows = df.shape[0]
                for row_idx in range(max_header_rows):
                    tr = soup.new_tag('tr')
                    col_idx = 0  # 列索引 (0-based)
                    while col_idx < df.shape[1]:
                        # 转换为 xlrd 的based 索引
                        cell_pos = (row_idx, col_idx)
                        if cell_pos in merge_info:
                            rowspan, colspan = merge_info[cell_pos]
                            cell_value = ws.cell_value(*cell_pos)
                            th = soup.new_tag('th', rowspan=str(rowspan), colspan=str(colspan))
                            th.string = str(cell_value)
                            tr.append(th)
                            col_idx += colspan
                        else:
                            if cell_pos not in merge_cell:
                                th = soup.new_tag('th')
                                th.string = str(df.iat[row_idx, col_idx])
                                tr.append(th)
                            col_idx += 1

                    thead.append(tr)
                table.append(thead)
                result.append(str(table.prettify()))

        return result

    @staticmethod
    def parse_csv(file_data: bytes) -> List[str]:
        if not isinstance(file_data, bytes):
            raise ValueError("file_data must be bytes")
        file_bio = BytesIO(file_data)
        df = pd.read_csv(file_bio, header=None).astype(str)

        # 有效性验证（防止全空数据）
        if df.map(lambda x: x.strip() if isinstance(x, str) else '').eq('').all().all():
            raise ValueError("Empty dataframe")

        md_table = df.to_markdown()

        # 添加表格元数据
        md_table = f"# Table: \n{md_table}"

        return [md_table]
